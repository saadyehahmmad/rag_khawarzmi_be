"""
Document loading and text splitting for ingestion.

Separates I/O (loaders) from vector-store logic so tests and tooling can import loaders alone.
"""

from __future__ import annotations

import re
from pathlib import Path

from langchain_community.document_loaders import Docx2txtLoader, PyPDFLoader, TextLoader
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

from core.text_ar import arabic_script_ratio
from ingestion.config import CHUNK_OVERLAP, CHUNK_SIZE

# Human-readable label for each system collection, used in heading breadcrumbs.
_SYSTEM_LABELS: dict[str, str] = {
    "designer": "Survey Designer",
    "admin": "Field Management",
    "callcenter": "Call Center",
    "runtime": "Survey Runtime",
}

# Matches markdown ATX headings: "## Heading Text"
_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+)$", re.MULTILINE)


def detect_language(text: str) -> str:
    """Return 'ar' when the Arabic script share exceeds 20 percent."""
    return "ar" if arabic_script_ratio(text) > 0.2 else "en"


def load_document(file_path: Path) -> list[Document]:
    """Load one file into LangChain Document objects; unsupported types return []."""
    ext = file_path.suffix.lower()
    try:
        if ext == ".pdf":
            loader = PyPDFLoader(str(file_path))
        elif ext == ".docx":
            loader = Docx2txtLoader(str(file_path))
        elif ext in (".txt", ".md"):
            loader = TextLoader(str(file_path), encoding="utf-8")
        elif ext == ".xlsx":
            return _load_xlsx_as_documents(file_path)
        else:
            print(f"  [SKIP] Unsupported format: {file_path.name}")
            return []
        docs = loader.load()
        print(f"  [OK] Loaded {len(docs)} segment(s) from {file_path.name}")
        return docs
    except Exception as exc:  # noqa: BLE001 — ingestion should continue across bad files
        print(f"  [ERROR] Failed to load {file_path.name}: {exc}")
        return []


def _load_xlsx_as_documents(file_path: Path) -> list[Document]:
    """Read first sheet of xlsx into one Document (simple cell text join)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  [SKIP] openpyxl not available for xlsx")
        return []
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb.active
    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) for c in row if c is not None and str(c).strip()]
        if cells:
            lines.append(" | ".join(cells))
    text = "\n".join(lines)
    wb.close()
    print(f"  [OK] Loaded xlsx rows from {file_path.name}")
    return [Document(page_content=text, metadata={"source": file_path.name})]


def _parse_heading_sections(text: str) -> list[tuple[str, str]]:
    """
    Split markdown text into (breadcrumb, section_text) pairs at every ATX heading.

    The breadcrumb is built from the active heading stack, e.g.:
        "3. Field Management System > 3.2 Survey Screen > 9. Tracking"

    Sections before the first heading (preamble) are returned with an empty breadcrumb
    and included only when they contain non-whitespace content.
    """
    matches = list(_HEADING_RE.finditer(text))
    if not matches:
        return [("", text)]

    sections: list[tuple[str, str]] = []
    # Stack entries: (level: int, title: str)
    heading_stack: list[tuple[int, str]] = []

    # Preamble — text before the first heading
    preamble = text[: matches[0].start()].strip()
    if preamble:
        sections.append(("", preamble))

    for i, match in enumerate(matches):
        level = len(match.group(1))  # number of '#' characters
        title = match.group(2).strip()

        # Pop all stack entries at the same or deeper level so the new heading replaces them.
        heading_stack = [(lvl, ttl) for lvl, ttl in heading_stack if lvl < level]
        heading_stack.append((level, title))

        breadcrumb = " > ".join(ttl for _, ttl in heading_stack)

        # Section body: from this heading to the start of the next one (or end of text).
        body_start = match.start()
        body_end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        section_text = text[body_start:body_end].strip()

        sections.append((breadcrumb, section_text))

    return sections


def enrich_with_section_context(docs: list[Document], system: str) -> list[Document]:
    """
    Prepend a heading breadcrumb context line to each markdown/text document section.

    For every document the function:
    1. Detects whether the source is a markdown or plain-text file.
    2. Parses the heading hierarchy and splits the text into sections.
    3. Prepends ``[{system_label} | {breadcrumb}]`` to each section's text so that
       downstream chunks produced by the splitter always carry the section context.
    4. Stores ``section_path`` and ``system_label`` in the document metadata.

    Non-markdown/non-text documents (PDF, DOCX, XLSX) are returned unchanged because
    their structure does not follow ATX-heading conventions.

    Returns a new list; the original documents are not mutated.
    """
    system_label = _SYSTEM_LABELS.get(system, system.title())
    enriched: list[Document] = []

    for doc in docs:
        source_file: str = doc.metadata.get("source_file", "")
        ext = Path(source_file).suffix.lower() if source_file else ""

        # Only enrich files whose heading structure we can reliably parse.
        if ext not in (".md", ".txt", ""):
            enriched.append(doc)
            continue

        text = doc.page_content or ""
        sections = _parse_heading_sections(text)

        if len(sections) <= 1 and not sections[0][0]:
            # No headings found — pass the document through with empty section_path.
            new_doc = Document(
                page_content=text,
                metadata={**doc.metadata, "section_path": "", "system_label": system_label},
            )
            enriched.append(new_doc)
            continue

        for breadcrumb, section_text in sections:
            if not section_text.strip():
                continue
            if breadcrumb:
                context_header = f"[{system_label} | {breadcrumb}]\n\n"
                enriched_text = context_header + section_text
            else:
                # Preamble before first heading — no breadcrumb, pass as-is.
                enriched_text = section_text

            new_doc = Document(
                page_content=enriched_text,
                metadata={
                    **doc.metadata,
                    "section_path": breadcrumb,
                    "system_label": system_label,
                },
            )
            enriched.append(new_doc)

    return enriched


def build_splitter() -> RecursiveCharacterTextSplitter:
    """Recursive splitter with Arabic-friendly separators (comma ، and newlines)."""
    return RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", ".", "،", " ", ""],
    )
